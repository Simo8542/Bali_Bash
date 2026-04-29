"""Genererer udlejning_case.xlsx - Excel beregner med justerbare inputs."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Farver (matcher repo-paletten)
INPUT_FILL  = PatternFill("solid", fgColor="FFF4D6")  # blød gul
AUTO_FILL   = PatternFill("solid", fgColor="E8F1E5")  # blød grøn
HEADER_FILL = PatternFill("solid", fgColor="1A1208")  # mørk
SUB_FILL    = PatternFill("solid", fgColor="C4622D")  # terracotta
RECOM_FILL  = PatternFill("solid", fgColor="D4A843")  # gold

WHITE = Font(color="FBF6EE", bold=True, name="Calibri", size=11)
GOLD  = Font(color="D4A843", bold=True)
BOLD  = Font(bold=True)
ITAL  = Font(italic=True, color="8A7D6B")

KR_FMT  = '#,##0" kr"'
PCT_FMT = '0.00"%"'
NUM_FMT = '#,##0'

thin = Side(border_style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Beregner"

# Kolonnebredder
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 12

def section(row, text, fill=HEADER_FILL, font=WHITE, span=6):
    """Skriv en farvet section-overskrift i hele bredden."""
    ws.cell(row=row, column=1, value=text).font = font
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, span+1):
        ws.cell(row=row, column=c).fill = fill
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)

def label(row, text, italic=False):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = ITAL if italic else Font()

def input_cell(row, col, value, fmt=KR_FMT):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = INPUT_FILL
    c.font = BOLD
    c.number_format = fmt
    c.border = BORDER
    return c

def auto_cell(row, col, formula, fmt=KR_FMT, bold=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = AUTO_FILL
    c.number_format = fmt
    if bold:
        c.font = BOLD
    return c

def unit(row, col, text):
    ws.cell(row=row, column=col, value=text).font = ITAL

# Title
ws.cell(row=1, column=1, value="UDLEJNING BUSINESS CASE").font = Font(bold=True, size=18, color="1A1208")
ws.cell(row=2, column=1, value="Helårsudlejning · Ceresbyen, Aarhus C").font = Font(italic=True, size=11, color="8A7D6B")
ws.row_dimensions[1].height = 28

# =================== INPUT ===================
section(4, "  INPUT  (juster gule felter)")

# Lån
section(6, "  LÅN  (Jyske Bank prioritetslån Cibor3)", fill=SUB_FILL)
label(7, "Restgæld");                input_cell(7,  2, 2086317);  unit(7,  3, "kr")
label(8, "Debitorrente");            input_cell(8,  2, 2.84, PCT_FMT); unit(8,  3, "% p.a.")
label(9, "Månedlig ydelse (før skat)"); input_cell(9, 2, 12150);   unit(9,  3, "kr/md")

# Ejendom
section(11, "  EJENDOM", fill=SUB_FILL)
label(12, "Offentlig ejendomsvurdering");  input_cell(12, 2, 4324000); unit(12, 3, "kr")
label(13, "Grundværdi");                   input_cell(13, 2, 2338000); unit(13, 3, "kr")
label(14, "Grundskyldspromille (Aarhus)"); input_cell(14, 2, 5.6, '0.00');  unit(14, 3, "‰")
label(15, "Kontant anskaffelsessum (KAO)"); input_cell(15, 2, 4875000); unit(15, 3, "kr")

# Drift
section(17, "  DRIFTSUDGIFTER  (månedlig)", fill=SUB_FILL)
label(18, "Ejerforeningsbidrag");          input_cell(18, 2, 1810); unit(18, 3, "kr/md")
label(19, "Indboforsikring");              input_cell(19, 2, 250);  unit(19, 3, "kr/md")
label(20, "Vedligehold (% af ejendomsværdi p.a.)"); input_cell(20, 2, 0.5, '0.0'); unit(20, 3, "%")
label(21, "Forbrug betalt af ejer (0 = lejer aconto)"); input_cell(21, 2, 0); unit(21, 3, "kr/md")
label(22, "Andre faste omkostninger");     input_cell(22, 2, 0);    unit(22, 3, "kr/md")

# Skat
section(24, "  SKATTESATSER", fill=SUB_FILL)
label(25, "Marg.skat positiv kapitalindkomst"); input_cell(25, 2, 37.7,  PCT_FMT); unit(25, 3, "%")
label(26, "Marg.skat negativ kapitalindkomst"); input_cell(26, 2, 25.6,  PCT_FMT); unit(26, 3, "%")
label(27, "Marg.skat personlig indkomst");      input_cell(27, 2, 52.07, PCT_FMT); unit(27, 3, "%")
label(28, "VSO foreløbig skat (opsparet)");      input_cell(28, 2, 22.0,  PCT_FMT); unit(28, 3, "%")
label(29, "KAO afkastsats");                     input_cell(29, 2, 3.0,   PCT_FMT); unit(29, 3, "%")

# Leje
section(31, "  LEJEINDTÆGT", fill=SUB_FILL)
label(32, "Månedlig leje (test-niveau)"); input_cell(32, 2, 16000); unit(32, 3, "kr/md")
label(33, "Tomgang (% af året)");          input_cell(33, 2, 0, PCT_FMT); unit(33, 3, "%")

# =================== AUTOBEREGNET ===================
# Cellereferencer (faste):
#   B7=restgaeld, B8=rente, B9=ydelse
#   B12=vurdering, B13=grundvaerdi, B14=promille, B15=anskaffelse
#   B18=ejerforening, B19=indbofors, B20=vedl%, B21=forbrug_ejer, B22=andre
#   B25=kap+, B26=kap-, B27=pers, B28=vso, B29=kao_sats
#   B32=leje, B33=tomgang

section(35, "  AUTOBEREGNET", fill=HEADER_FILL)

label(36, "Månedlige renter")
auto_cell(36, 2, "=B7*B8/100/12"); unit(36, 3, "kr/md")

label(37, "Månedligt afdrag")
auto_cell(37, 2, "=B9-B36"); unit(37, 3, "kr/md")

label(38, "Årlige renter")
auto_cell(38, 2, "=B36*12"); unit(38, 3, "kr/år")

label(39, "Årligt afdrag")
auto_cell(39, 2, "=B37*12"); unit(39, 3, "kr/år")

label(40, "Grundskyld månedlig")
auto_cell(40, 2, "=B13*B14/1000/12"); unit(40, 3, "kr/md")

label(41, "Vedligehold månedlig")
auto_cell(41, 2, "=B12*B20/100/12"); unit(41, 3, "kr/md")

label(42, "Drift TOTAL månedlig")
auto_cell(42, 2, "=B18+B19+B40+B41+B21+B22", bold=True); unit(42, 3, "kr/md")

label(43, "Drift TOTAL årlig")
auto_cell(43, 2, "=B42*12", bold=True); unit(43, 3, "kr/år")

label(44, "Effektiv leje månedlig (efter tomgang)")
auto_cell(44, 2, "=B32*(1-B33/100)"); unit(44, 3, "kr/md")

label(45, "Effektiv leje årlig")
auto_cell(45, 2, "=B44*12"); unit(45, 3, "kr/år")

# =================== CASHFLOW ===================
section(47, "  CASHFLOW VED VALGTE LEJE  (kr/md, før skat)")

label(48, "+ Bruttoleje");   auto_cell(48, 2, "=B44")
label(49, "− Drift");        auto_cell(49, 2, "=-B42")
label(50, "− Ydelse (rente+afdrag)"); auto_cell(50, 2, "=-B9")

label(51, "= LIKVIDITET (før skat)");
c = auto_cell(51, 2, "=B48+B49+B50", bold=True)
c.fill = RECOM_FILL

label(52, "+ Afdrag tilbageført som formueopbygning")
auto_cell(52, 2, "=B37")

label(53, "= ØKONOMISK OVERSKUD (før skat)")
c = auto_cell(53, 2, "=B51+B52", bold=True)
c.fill = RECOM_FILL

# Årligt skattepligtigt resultat
section(55, "  SKATTEPLIGTIGT RESULTAT  (årligt, regnskabsmæssigt)")

label(56, "+ Bruttoleje (årlig)");      auto_cell(56, 2, "=B45")
label(57, "− Driftsudgifter");           auto_cell(57, 2, "=-B43")
label(58, "= Resultat FØR renter");      auto_cell(58, 2, "=B56+B57", bold=True)
label(59, "− Renter");                   auto_cell(59, 2, "=-B38")
label(60, "= Resultat EFTER renter");    auto_cell(60, 2, "=B58+B59", bold=True)

ws.cell(row=61, column=1, value="Note: ejendomsværdiskat = 0 ved 100% helårsudlejning hele året").font = ITAL

# =================== SKATTEMETODER ===================
section(63, "  SKATTEMETODER · SAMMENLIGNING  (årligt)")

# Header-række med metodenavne
methods = [
    (2, "M1  Personlig\nkapitalindkomst"),
    (3, "M2  VSO\n(udbetalt)"),
    (4, "M3  VSO\n(opsparet, foreløbig)"),
    (5, "M4  KAO\n(kapitalafkastordning)"),
]
ws.cell(row=64, column=1, value="").fill = HEADER_FILL
for col, name in methods:
    c = ws.cell(row=64, column=col, value=name)
    c.font = WHITE
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[64].height = 36

# Linje 65: Resultat før renter
label(65, "Resultat før renter (grundlag)")
for col, _ in methods:
    auto_cell(65, col, "=B58")

# Linje 66: Renter trukket fra grundlaget (kun M2/M3/M4)
label(66, "− Renter (i grundlaget)")
auto_cell(66, 2, 0)
auto_cell(66, 3, "=-B38")
auto_cell(66, 4, "=-B38")
auto_cell(66, 5, "=-B38")

# Linje 67: KAO kapitalafkast (flyttes til kap.indk.)
label(67, "− KAO kapitalafkast (flyttes til kap.indk.)")
auto_cell(67, 2, 0)
auto_cell(67, 3, 0)
auto_cell(67, 4, 0)
# Kapitalafkast = MIN(anskaffelse × sats, MAX(0, resultat efter renter))
auto_cell(67, 5, "=-MIN($B$15*$B$29/100, MAX(0, B58-B38))")

# Linje 68: Skattegrundlag (personlig del / kap.indk.del for M1)
label(68, "= Skattegrundlag")
for col, _ in methods:
    auto_cell(68, col, f"=SUM({get_column_letter(col)}65:{get_column_letter(col)}67)", bold=True)

# Linje 69: Skat af grundlaget
label(69, "Skat af grundlag")
# M1: pos kap.indk. hvis >=0 ellers neg kap.indk.
auto_cell(69, 2, "=IF(B68>=0, B68*$B$25/100, B68*$B$26/100)")
# M2 VSO udbetalt: personlig (~52%)
auto_cell(69, 3, "=C68*$B$27/100")
# M3 VSO opsparet: foreløbig 22%
auto_cell(69, 4, "=D68*$B$28/100")
# M4 KAO: personlig del × pers + kapitalafkast × kap+
auto_cell(69, 5, "=E68*$B$27/100 + MIN($B$15*$B$29/100, MAX(0, B58-B38))*$B$25/100")

# Linje 70: − værdi af privat rentefradrag (kun M1)
label(70, "− Værdi af privat rentefradrag")
auto_cell(70, 2, "=-B38*$B$26/100")
auto_cell(70, 3, 0)  # rentefradrag allerede i grundlaget
auto_cell(70, 4, 0)
auto_cell(70, 5, 0)

# Linje 71: NETTOSKAT (årlig)
label(71, "= NETTOSKAT (årlig)")
for col, _ in methods:
    c = auto_cell(71, col, f"={get_column_letter(col)}69+{get_column_letter(col)}70", bold=True)
    c.fill = RECOM_FILL

# Linje 72: Overskud efter skat (årligt, ekskl. afdrag = formue)
label(72, "Overskud efter skat (årligt, formue ekskl.)")
for col, _ in methods:
    auto_cell(72, col, f"=B60-{get_column_letter(col)}71")

# Linje 73: Likviditet efter skat (årligt, INKL. afdrag som cash-out)
label(73, "Likviditet efter skat (årligt, formue inkl.)")
for col, _ in methods:
    auto_cell(73, col, f"={get_column_letter(col)}72-B39")

# Linje 74: Likviditet efter skat (månedlig)
label(74, "Likviditet efter skat (kr/md)")
for col, _ in methods:
    c = auto_cell(74, col, f"={get_column_letter(col)}73/12", bold=True)

# Linje 75: Reelt månedligt overskud (formue medregnet)
label(75, "Reelt overskud m. formueopbygning (kr/md)")
for col, _ in methods:
    c = auto_cell(75, col, f"={get_column_letter(col)}72/12", bold=True)
    c.fill = RECOM_FILL

# =================== ANBEFALING ===================
section(77, "  ANBEFALING")

label(78, "Laveste reelle skat (ekskl. M3 VSO opsp. der kun udskyder)")
# Find min af B71, C71, E71 (M3 udelukket) og match med metodenavn
auto_cell(78, 2,
    '=IF(B71=MIN(B71,C71,E71),"M1 Kapitalindkomst",'
    'IF(C71=MIN(B71,C71,E71),"M2 VSO udbetalt","M4 KAO"))',
    fmt='@', bold=True
)
ws.cell(row=78, column=2).fill = RECOM_FILL

label(79, "Skat ved anbefalet metode (årlig)")
auto_cell(79, 2, '=MIN(B71,C71,E71)', bold=True)

label(80, "Reelt overskud per md (anbefalet)")
auto_cell(80, 2, '=(B60-MIN(B71,C71,E71))/12', bold=True)

ws.cell(row=81, column=1, value=("M3 (VSO opsparet) sparer mere på kort sigt men kun fordi skatten "
        "udskydes — ved hævning betales op til personlig sats.")).font = ITAL
ws.merge_cells(start_row=81, start_column=1, end_row=81, end_column=6)

# =================== BREAK-EVEN ===================
section(83, "  BREAK-EVEN MÅNEDLIG LEJE  (kr/md, skat indregnet)")

# Header
ws.cell(row=84, column=1, value="").fill = HEADER_FILL
hdr = [(2, "M1 Kap.indk."), (3, "M2 VSO udb."), (4, "M3 VSO opsp."), (5, "M4 KAO")]
for col, name in hdr:
    c = ws.cell(row=84, column=col, value=name)
    c.font = WHITE; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

# Break-even formel (lukket form per metode):
# Likviditet=0:  leje*12 - drift_a - ydelse_a - skat(leje) = 0
# Økonomisk=0:   leje*12 - drift_a - renter_a - skat(leje) = 0
# Skat afhænger lineært af lejen, så vi løser direkte.
# Lad L = årlig leje. Resultat før renter = L - drift.
# Skat per metode (positivt grundlag antages):
#   M1:  (L-drift)*kap+/100  -  renter*kap-/100
#   M2:  (L-drift-renter)*pers/100
#   M3:  (L-drift-renter)*vso/100
#   M4:  (L-drift-renter-min(kapafkast, L-drift-renter))*pers/100
#         + min(kapafkast, L-drift-renter)*kap+/100
#         (når L-drift-renter < kapafkast: alt til kap+, dvs (L-drift-renter)*kap+/100)
#
# For likviditet break-even: L = drift_a + ydelse_a + skat(L)
# Vi opstiller eksplicit formel for hver metode.

# Hjælpevariable som navngivne references:
#   drift_a = B43, ydelse_a = B9*12, renter_a = B38, kapafkast = B15*B29/100
#   kap+ = B25/100, kap- = B26/100, pers = B27/100, vso = B28/100

# M1: L*kap+ - drift_a*kap+ - renter_a*kap- = ydelse_a + drift_a - L
# → L*(1+kap+) = ydelse_a + drift_a + drift_a*kap+ + renter_a*kap-
# Hvis drift_a*kap+ flyttes: L*(1+kap+) = ydelse_a + drift_a*(1+kap+) + renter_a*kap-
# → L = drift_a + (ydelse_a + renter_a*kap-) / (1+kap+) ... NEJ omregn fra start.
#
# Likviditet=0 betyder årlig: L*12 - drift_a - 12*ydelse_md - skat = 0
# (L her er årlig)
# Lad mig bruge L = årlig leje:
# 0 = L - drift_a - ydelse_a - skat
# skat_M1 = (L - drift_a)*kp - renter_a*km   (forudsat L-drift_a >= 0)
# 0 = L - drift_a - ydelse_a - (L-drift_a)*kp + renter_a*km
# L*(1-kp) = drift_a + ydelse_a - drift_a*kp - renter_a*km
# L*(1-kp) = drift_a*(1-kp) + ydelse_a - renter_a*km
# L = drift_a + (ydelse_a - renter_a*km)/(1-kp)
# årlig → /12 for månedlig

# M2: skat = (L-drift_a-renter_a)*pers
# 0 = L - drift_a - ydelse_a - (L-drift_a-renter_a)*pers
# L*(1-pers) = drift_a + ydelse_a - drift_a*pers - renter_a*pers
# L*(1-pers) = (drift_a+renter_a)*(1-pers) + ydelse_a - renter_a   ... lad mig omregne:
# L*(1-pers) = drift_a + ydelse_a - drift_a*pers - renter_a*pers
# L = [drift_a*(1-pers) + ydelse_a - renter_a*pers] / (1-pers)
# L = drift_a + (ydelse_a - renter_a*pers)/(1-pers)

# M3 samme form med vso i stedet for pers

# M4: hvis kapafkast >= L-drift_a-renter_a (typisk for store anskaffelsessummer):
#  skat = (L-drift_a-renter_a)*kp+
# Samme form som M1 men med kp+ og uden separat rentefradrag:
# 0 = L - drift_a - ydelse_a - (L-drift_a-renter_a)*kp+
# L*(1-kp+) = drift_a + ydelse_a - drift_a*kp+ - renter_a*kp+
# L = drift_a + (ydelse_a - renter_a*kp+)/(1-kp+) - og minus den rest...
# Faktisk: L*(1-kp+) = (drift_a+renter_a)*(1-kp+) + ydelse_a - renter_a + renter_a*kp+...
# Lad mig være præcis:
# L - drift_a*(1) - ydelse_a*(1) = (L-drift_a-renter_a)*kp+
# L*(1-kp+) = drift_a + ydelse_a - drift_a*kp+ - renter_a*kp+
# L = (drift_a*(1-kp+) + ydelse_a - renter_a*kp+) / (1-kp+)
# L = drift_a + (ydelse_a - renter_a*kp+)/(1-kp+)

# LIKVIDITET BREAK-EVEN (årlig L, divideres med 12)
label(85, "Lejen for likviditet = 0 (afdrag betales af lejen)")

# M1
auto_cell(85, 2, "=(B43 + (B9*12 - B38*B26/100)/(1-B25/100))/12", bold=True)
# M2
auto_cell(85, 3, "=(B43 + (B9*12 - B38*B27/100)/(1-B27/100))/12", bold=True)
# M3 (VSO opsp.)
auto_cell(85, 4, "=(B43 + (B9*12 - B38*B28/100)/(1-B28/100))/12", bold=True)
# M4 (antagelse: kapafkast > resultat → alt til kap+)
auto_cell(85, 5, "=(B43 + (B9*12 - B38*B25/100)/(1-B25/100))/12", bold=True)

# ØKONOMISK BREAK-EVEN: 0 = L - drift_a - renter_a - skat
# (afdrag tæller IKKE som udgift - er formueopbygning)
label(86, "Lejen for økonomisk overskud = 0 (afdrag = formue)")

# M1: 0 = L - drift_a - renter_a - (L-drift_a)*kp + renter_a*km
# L*(1-kp) = drift_a*(1-kp) + renter_a*(1-km)
# L = drift_a + renter_a*(1-km)/(1-kp)
auto_cell(86, 2, "=(B43 + B38*(1-B26/100)/(1-B25/100))/12", bold=True)

# M2: 0 = L - drift_a - renter_a - (L-drift_a-renter_a)*pers
# (L-drift_a-renter_a)*(1-pers) = 0  → L = drift_a + renter_a
auto_cell(86, 3, "=(B43 + B38)/12", bold=True)

# M3: samme som M2: L = drift_a + renter_a
auto_cell(86, 4, "=(B43 + B38)/12", bold=True)

# M4: 0 = L - drift_a - renter_a - (L-drift_a-renter_a)*kp+
# samme form som M2 men med kp+: → L = drift_a + renter_a (når skat-grundlag = 0 → 0 skat)
auto_cell(86, 5, "=(B43 + B38)/12", bold=True)

# Forklaring
ws.cell(row=87, column=1, value="Likviditet=0: lejen dækker drift + hele ydelsen (incl. afdrag). Du lægger 0 til.").font = ITAL
ws.cell(row=88, column=1, value="Økon.=0: lejen dækker drift + renter. Afdrag = formueopbygning oveni.").font = ITAL

# =================== FORKLARINGER FANE ===================
ws2 = wb.create_sheet("Forklaringer")
ws2.column_dimensions['A'].width = 100

forklaringer = [
    ("FARVEKODER", True),
    ("Gule felter:    Justerbare INPUTS - skriv dine egne tal.", False),
    ("Grønne felter:  Auto-beregnede outputs (formler).", False),
    ("Mørke felter:   Sektionsoverskrifter.", False),
    ("Gylden:         Vigtige nøgletal og anbefaling.", False),
    ("", False),
    ("METODER", True),
    ("M1 Personlig kapitalindkomst (default ved fuld helårsudlejning):", False),
    ("    Lejeoverskud (før renter) beskattes som positiv kap.indk. (~37,7%).", False),
    ("    Renter er privat negativ kap.indk. - fradragsværdi ~25,6%.", False),
    ("    Ingen særlig ordning kræves.", False),
    ("", False),
    ("M2 VSO udbetalt: hele virksomhedsoverskuddet hæves som personlig indkomst.", False),
    ("    Renter flyttes til virksomhedens resultat → personlig indk. fradrag (~52%).", False),
    ("    Resultat efter renter beskattes med ~52% (kan udløse topskat).", False),
    ("    Kræver fuldt adskilt regnskab og bankkonto.", False),
    ("", False),
    ("M3 VSO opsparet: overskud beholdes i virksomheden, foreløbig 22% skat.", False),
    ("    Ved senere hævning betales differencen op til personlig sats.", False),
    ("    Likviditetsforbedring nu - skat udskudt, ikke fjernet.", False),
    ("", False),
    ("M4 KAO Kapitalafkastordningen: kapitalafkast (3% af anskaffelsessum) flyttes", False),
    ("    fra personlig indkomst til kapitalindkomst (37,7% i stedet for 52%).", False),
    ("    Ved store anskaffelsessummer (som dit 4,875 mio) kan hele resultatet flyttes.", False),
    ("    Forenklet beregning - ignorerer detail-reguleringer i KAO-regelsættet.", False),
    ("", False),
    ("FORBEHOLD", True),
    ("- Ejendomsværdiskat = 0 ved 100% helårsudlejning hele året.", False),
    ("- Ved erhvervsmæssig udlejning kan parcelhusreglen mistes → evt. skat ved salg.", False),
    ("- Lejen skal være markedsleje. Skat kan fiksere lejen ved underleje.", False),
    ("- Cibor3-renten varierer kvartalsvist - juster B8 ved rentetilpasning.", False),
    ("- Vedligeholdsfradrag dækker kun reparationer, ikke forbedringer.", False),
    ("- Konsulter en revisor før valg af VSO/KAO i selvangivelsen.", False),
    ("", False),
    ("BREAK-EVEN", True),
    ("Likviditet = 0:", False),
    ("    Lejen dækker drift + hele ydelsen (renter + afdrag) + skat.", False),
    ("    Du tilbageholder 0 kr af egen indkomst per måned.", False),
    ("Økonomisk overskud = 0:", False),
    ("    Lejen dækker drift + renter + skat.", False),
    ("    Afdragene (~7.200 kr/md) er ren formueopbygning - ikke en omkostning.", False),
    ("    Reelt overskud per md = afdrag = formuetilvækst.", False),
]

for i, (text, is_header) in enumerate(forklaringer, start=1):
    cell = ws2.cell(row=i, column=1, value=text)
    if is_header:
        cell.font = Font(bold=True, size=12, color="C4622D")
        cell.fill = PatternFill("solid", fgColor="FFF4D6")

# Gem fil
out = "/home/user/Bali_Bash/udlejning_case.xlsx"
wb.save(out)
print(f"Gemt: {out}")
